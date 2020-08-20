import requests
import json
import jsonpath
import openpyxl

class common :
    def __init__(self,filepath,sheetname):
        global wk
        global sh
        wk = openpyxl.load_workbook(filepath)
        sh = wk[sheetname]

    def fetch_raw_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_name(self):
        c = sh.max_column
        li=[]
        for i in range (1,c+1):
            cell =sh.cell(row= 1,column= i)
            li.insert(i-1,cell.value)
            return li

    def update_request_with_data(self,rownumber,jsonrequest,keylist):
        c = sh.max_column
        for i in range (1,c+1):
            cell= sh.cell(row=rownumber,column=i)
            jsonrequest[keylist[i-1]] = cell.value
            return jsonrequest
