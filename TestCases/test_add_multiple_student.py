import requests
import json
import jsonpath
import openpyxl
def test_add_multiple_student_data():
    #api code
    api_url = "http://www.thetestingworldapi.com/api/studentsDetails"
    f = open("D://api_json/py.txt", 'r')
    json_request = json.loads(f.read())
    #print(json_request)
   # print(json_request['first_name'])

    #excel code
    wk = openpyxl.load_workbook("D://api_json/xl_data.xlsx") # load work book
    sh = wk['Sheet1']                                        # load sheet
    rows = sh.max_row   # load maximum rows in sheet


    for i in range (2,rows+1) :
        first_name = sh.cell(row=i,column =1)   #create variable for cell
        middle_name = sh.cell(row=i,column = 2)
        last_name = sh.cell(row=i,column = 3)
        d_o_b = sh.cell(row=i,column =4)

        # to read data from cell like first_name.value
        #put that value in json values
        json_request['first_name'] = first_name.value
        json_request['middle_name'] = middle_name.value
        json_request['last_name'] = last_name.value
        json_request['date_of_birth'] = d_o_b.value

        # send the request to api
        response = requests.post(api_url,json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201